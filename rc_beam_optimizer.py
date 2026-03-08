#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import math
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Unit conversion
KN_PER_KG = 9.80665e-3

BAR = {
    # label: (nominal_diameter_mm, area_mm2, unit_weight_kg_per_m)
    "D10": (10.0, 71.0, 0.560),
    "D13": (13.0, 127.0, 0.995),
    "D16": (16.0, 199.0, 1.560),
    "D19": (19.0, 287.0, 2.250),
    "D22": (22.0, 387.0, 3.040),
    "D25": (25.0, 507.0, 3.980),
    "D29": (29.0, 642.0, 5.040),
    "D32": (32.0, 794.0, 6.230),
    "D35": (35.0, 957.0, 7.510),
}


def norm(v):
    return "" if v is None else str(v).strip().lower().replace(" ", "")


def bval(v, d=False):
    if v is None:
        return d
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in {"1", "true", "t", "yes", "y", "on"}:
        return True
    if s in {"0", "false", "f", "no", "n", "off"}:
        return False
    return d


def fval(v, d=0.0):
    if v is None:
        return d
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return d
    try:
        return float(s)
    except Exception:
        return d


def ival(v, d=0):
    if v is None:
        return d
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(round(v))
    try:
        return int(round(float(str(v).strip())))
    except Exception:
        return d


def clamp(v, lo, hi):
    return max(lo, min(hi, v))


def kv(ws):
    d = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        k = norm(r[0] if len(r) > 0 else None)
        if k:
            d[k] = r[1] if len(r) > 1 else None
    return d


def header_map(ws):
    m = {}
    for c, cell in enumerate(ws[1], start=1):
        k = norm(cell.value)
        if k:
            m[k] = c - 1
    return m


def pick(row, m, *keys):
    for k in keys:
        i = m.get(norm(k))
        if i is not None and i < len(row):
            return row[i]
    return None


def req_sheet(wb, name):
    if name not in wb.sheetnames:
        raise ValueError(f"Missing sheet: {name}")
    return wb[name]


def bar(lbl):
    k = str(lbl).strip().upper()
    if k not in BAR:
        raise ValueError(f"Unknown bar: {lbl}")
    d_mm, a_mm2, w_kg_m = BAR[k]
    return d_mm, a_mm2, w_kg_m


def read_input(path):
    wb = load_workbook(path, data_only=True)
    ws_m = req_sheet(wb, "MATERIALS")
    ws_s = req_sheet(wb, "SETTINGS")
    ws_c = req_sheet(wb, "COST")
    ws_b = req_sheet(wb, "BEAMS")
    ws_k = req_sheet(wb, "CANDIDATES")

    mk, sk, ck = kv(ws_m), kv(ws_s), kv(ws_c)
    mat = {
        # Stress unit: N/mm2 (same numeric as MPa)
        "fc": fval(mk.get("fc_n_mm2"), fval(mk.get("fc_mpa"), 30.0)),
        "fy": fval(mk.get("fy_main_n_mm2"), fval(mk.get("fy_main_mpa"), 490.0)),
        "fyv": fval(mk.get("fy_shear_n_mm2"), fval(mk.get("fy_shear_mpa"), 295.0)),
        "cover": fval(mk.get("cover_mm"), 40.0),
        "gamma": fval(mk.get("concrete_unit_weight_kn_m3"), 24.0),
        "phi_m": fval(mk.get("phi_flexure"), 0.9),
        "phi_v": fval(mk.get("phi_shear"), 0.75),
        "rho_max": fval(mk.get("rho_max"), 0.025),
        "fs_lim": fval(mk.get("steel_stress_limit_n_mm2"), fval(mk.get("steel_stress_limit_mpa"), 280.0)),
        "defl_ratio": fval(mk.get("deflection_limit_ratio"), 250.0),
        "ei_fac": fval(mk.get("effective_stiffness_factor"), 0.35),
        "s_clear_min": fval(mk.get("min_clear_spacing_mm"), 25.0),
        "n_div": max(80, ival(mk.get("default_n_div"), 400)),
    }
    setts = {
        "gD": fval(sk.get("load_factor_dead"), 1.2),
        "gL": fval(sk.get("load_factor_live"), 1.6),
        "mode": str(sk.get("objective_mode") or "HYBRID").strip().upper(),
        "w_cost": fval(sk.get("cost_weight"), 1.0),
        "w_co2": fval(sk.get("co2_weight"), 0.0),
        "max_rows": max(100, ival(sk.get("max_check_rows"), 20000)),
        "all_checks": bval(sk.get("output_all_checks"), True),
    }
    cost_steel_jpy_kg = fval(ck.get("rebar_jpy_kg"), 0.0)
    if cost_steel_jpy_kg <= 0.0:
        # backward compatibility for old input key
        cost_steel_jpy_kg = fval(ck.get("rebar_jpy_kn"), 160.0 / KN_PER_KG) * KN_PER_KG

    co2_steel_kgco2_kg = fval(ck.get("rebar_co2_kg_kg"), 0.0)
    if co2_steel_kgco2_kg <= 0.0:
        # backward compatibility for old input key
        co2_steel_kgco2_kg = fval(ck.get("rebar_co2_kgco2_kn"), 1.4 / KN_PER_KG) * KN_PER_KG

    cost = {
        "c_conc": fval(ck.get("concrete_jpy_m3"), 18000.0),
        "c_steel": cost_steel_jpy_kg,
        "c_form": fval(ck.get("formwork_jpy_m2"), 4200.0),
        "e_conc": fval(ck.get("concrete_co2_kg_m3"), 320.0),
        "e_steel": co2_steel_kgco2_kg,
        "e_form": fval(ck.get("formwork_co2_kg_m2"), 8.0),
    }

    warnings = []
    beams, cands = [], []

    mb = header_map(ws_b)
    for rix, row in enumerate(ws_b.iter_rows(min_row=2, values_only=True), start=2):
        if not bval(pick(row, mb, "use"), False):
            continue
        bid = str(pick(row, mb, "beamid", "beam_id") or "").strip()
        if not bid:
            warnings.append(f"BEAMS row {rix}: BeamID blank")
            continue
        span = fval(pick(row, mb, "span_m", "span"), 0.0)
        trib = fval(pick(row, mb, "tributarywidth_m", "tributary_width_m"), 0.0)
        if span <= 0:
            warnings.append(f"BEAMS row {rix} {bid}: span<=0 skipped")
            continue
        beams.append({
            "id": bid,
            "span": span,
            "trib": max(0.0, trib),
            "qD": fval(pick(row, mb, "deadload_kn_m2", "dead_load_kn_m2"), 0.0),
            "qL": fval(pick(row, mb, "liveload_kn_m2", "live_load_kn_m2"), 0.0),
            "PD": fval(pick(row, mb, "pointdead_kn", "point_dead_kn"), 0.0),
            "PL": fval(pick(row, mb, "pointlive_kn", "point_live_kn"), 0.0),
            "r": clamp(fval(pick(row, mb, "pointposratio", "point_pos_ratio"), 0.5), 0.0, 1.0),
        })

    mc = header_map(ws_k)
    for rix, row in enumerate(ws_k.iter_rows(min_row=2, values_only=True), start=2):
        if not bval(pick(row, mc, "use"), False):
            continue
        rank = ival(pick(row, mc, "rank"), 0)
        if rank <= 0:
            warnings.append(f"CANDIDATES row {rix}: rank<=0")
            continue
        sec = str(pick(row, mc, "sectionid", "section_id") or f"R{rank}").strip()
        c = {
            "rank": rank,
            "sec": sec,
            "b": fval(pick(row, mc, "width_mm", "b_mm"), 0.0),
            "h": fval(pick(row, mc, "depth_mm", "h_mm"), 0.0),
            "nb": max(1, ival(pick(row, mc, "bottombars_n", "bottom_bars_n"), 0)),
            "db": str(pick(row, mc, "bottombar", "bottom_bar") or "").strip().upper(),
            "nt": max(0, ival(pick(row, mc, "topbars_n", "top_bars_n"), 0)),
            "dt": str(pick(row, mc, "topbar", "top_bar") or "").strip().upper(),
            "legs": max(0, ival(pick(row, mc, "stirruplegs_n", "stirrup_legs_n"), 2)),
            "ds": str(pick(row, mc, "stirrupbar", "stirrup_bar") or "").strip().upper(),
            "s": fval(pick(row, mc, "stirrupspacing_mm", "stirrup_spacing_mm"), 150.0),
        }
        if c["legs"] > 0:
            c["s"] = max(50.0, c["s"])
        else:
            c["s"] = 0.0
            c["ds"] = ""

        if c["b"] <= 0 or c["h"] <= 0 or not c["db"]:
            warnings.append(f"CANDIDATES row {rix} {sec}: missing geometry/bar")
            continue
        if c["nt"] > 0 and not c["dt"]:
            warnings.append(f"CANDIDATES row {rix} {sec}: top bars set but top bar type blank")
            continue
        if c["legs"] > 0 and not c["ds"]:
            warnings.append(f"CANDIDATES row {rix} {sec}: stirrup legs set but stirrup bar type blank")
            continue
        try:
            bar(c["db"])
            if c["legs"] > 0:
                bar(c["ds"])
            if c["nt"] > 0:
                bar(c["dt"])
        except ValueError as e:
            warnings.append(f"CANDIDATES row {rix} {sec}: {e}")
            continue
        cands.append(c)

    if not beams:
        raise ValueError("No active rows in BEAMS")
    if not cands:
        raise ValueError("No active rows in CANDIDATES")
    cands.sort(key=lambda x: (x["rank"], x["sec"]))
    return mat, setts, cost, beams, cands, warnings


def support_reactions(span, w, p, r):
    L = max(0.0, float(span))
    a = (0.5 if p == 0 else clamp(r, 0.0, 1.0)) * L
    if L <= 0.0:
        return a, 0.0, 0.0
    R1 = w * L * 0.5 + p * (L - a) / L
    R2 = w * L * 0.5 + p * a / L
    return a, R1, R2


def env(span, w, p, r, n):
    L = span
    a, R1, R2 = support_reactions(span, w, p, r)
    Mx, Vx, xM, xV = 0.0, max(abs(R1), abs(R2)), 0.0, (0.0 if abs(R1) >= abs(R2) else L)
    n = max(40, n)
    for i in range(n + 1):
        x = L * i / n
        M = R1 * x - 0.5 * w * x * x - p * max(0.0, x - a)
        V = R1 - w * x - (p if x >= a else 0.0)
        if abs(M) > Mx:
            Mx, xM = abs(M), x
        if abs(V) > Vx:
            Vx, xV = abs(V), x
    return Mx, xM, Vx, xV


def defl_mm(span, w, p, r, ei, n):
    if span <= 0 or ei <= 0:
        return float("inf"), 0.0
    L = span * 1000.0
    a = (0.5 if p == 0 else clamp(r, 0.0, 1.0)) * L
    wN, PN = w, p * 1000.0
    R1 = wN * L * 0.5 + PN * (L - a) / L
    n = max(80, n)
    xs = [L * i / n for i in range(n + 1)]
    k = []
    for x in xs:
        M = R1 * x - 0.5 * wN * x * x - PN * max(0.0, x - a)
        k.append(M / ei)
    i1 = [0.0]
    for i in range(1, len(xs)):
        dx = xs[i] - xs[i - 1]
        i1.append(i1[-1] + 0.5 * (k[i - 1] + k[i]) * dx)
    i2 = [0.0]
    for i in range(1, len(xs)):
        dx = xs[i] - xs[i - 1]
        i2.append(i2[-1] + 0.5 * (i1[i - 1] + i1[i]) * dx)
    c1 = -i2[-1] / L
    y = [i2[i] + c1 * xs[i] for i in range(len(xs))]
    ym, xm = 0.0, 0.0
    for i, yy in enumerate(y):
        if abs(yy) > ym:
            ym, xm = abs(yy), xs[i] / 1000.0
    return ym, xm


def response_arrays(span, w_u, p_u, w_s, p_s, r, ei, n=20):
    if span <= 0:
        return [], [], [], []

    n = max(4, int(n))
    x_u = [span * i / n for i in range(n + 1)]
    a_u, R1_u, _R2_u = support_reactions(span, w_u, p_u, r)
    V_u, M_u = [], []
    for x in x_u:
        M_u.append(R1_u * x - 0.5 * w_u * x * x - p_u * max(0.0, x - a_u))
        V_u.append(R1_u - w_u * x - (p_u if x >= a_u else 0.0))

    if ei <= 0:
        return x_u, V_u, M_u, [None] * len(x_u)

    L = span * 1000.0
    x_s = [L * i / n for i in range(n + 1)]
    a_s = (0.5 if p_s == 0 else clamp(r, 0.0, 1.0)) * L
    wN = w_s
    PN = p_s * 1000.0
    R1_s = wN * L * 0.5 + PN * (L - a_s) / L

    curv = []
    for x in x_s:
        M = R1_s * x - 0.5 * wN * x * x - PN * max(0.0, x - a_s)
        curv.append(M / ei)

    i1 = [0.0]
    for i in range(1, len(x_s)):
        dx = x_s[i] - x_s[i - 1]
        i1.append(i1[-1] + 0.5 * (curv[i - 1] + curv[i]) * dx)

    i2 = [0.0]
    for i in range(1, len(x_s)):
        dx = x_s[i] - x_s[i - 1]
        i2.append(i2[-1] + 0.5 * (i1[i - 1] + i1[i]) * dx)

    c1 = -i2[-1] / L
    y_mm = [i2[i] + c1 * x_s[i] for i in range(len(x_s))]
    return x_u, V_u, M_u, y_mm

def eval_cand(beam, cand, mat, setts, cost, n_override=0):
    b, h = cand["b"], cand["h"]
    db, Ab, wb = bar(cand["db"])
    has_stirrups = cand["legs"] > 0
    if has_stirrups:
        ds, As1, ws = bar(cand["ds"])
    else:
        ds, As1, ws = 0.0, 0.0, 0.0
    if cand["nt"] > 0:
        dt, _At, wt = bar(cand["dt"])
    else:
        dt, wt = 0.0, 0.0

    n = n_override if n_override > 0 else mat["n_div"]
    sw = mat["gamma"] * (b / 1000.0) * (h / 1000.0)
    wD = beam["qD"] * beam["trib"] + sw
    wL = beam["qL"] * beam["trib"]
    wS, pS = wD + wL, beam["PD"] + beam["PL"]
    wU = setts["gD"] * wD + setts["gL"] * wL
    pU = setts["gD"] * beam["PD"] + setts["gL"] * beam["PL"]

    Mu, xmu, Vu, xvu = env(beam["span"], wU, pU, beam["r"], n)
    Ms, xms, _vs, _xvs = env(beam["span"], wS, pS, beam["r"], n)
    a_p, R1s, R2s = support_reactions(beam["span"], wS, pS, beam["r"])
    _a_u, R1u, R2u = support_reactions(beam["span"], wU, pU, beam["r"])

    As = cand["nb"] * Ab
    Asv = cand["legs"] * As1
    Ec = 4700.0 * math.sqrt(max(mat["fc"], 0.0))
    Ig = b * h ** 3 / 12.0
    ei = Ec * mat["ei_fac"] * Ig
    d = h - mat["cover"] - ds - 0.5 * db
    dallow = beam["span"] * 1000.0 / max(mat["defl_ratio"], 1.0)
    dmm, xd = defl_mm(beam["span"], wS, pS, beam["r"], ei, n)

    req_clear = max(mat["s_clear_min"], db)
    if cand["nb"] <= 1:
        clear, okC, uC = None, True, 0.0
    else:
        net = b - 2.0 * (mat["cover"] + ds)
        clear = (net - cand["nb"] * db) / (cand["nb"] - 1)
        okC = clear >= req_clear
        uC = req_clear / clear if clear and clear > 0 else float("inf")

    if d <= 0:
        Mn = phiMn = Vc = Vs = phiVn = 0.0
        okM = okV = okD = okFs = okR = False
        okS = not has_stirrups
        uM = uV = uD = uFs = uR = float("inf")
        s_lim = 0.0
        uS = 0.0 if okS else float("inf")
        rho = float("inf")
        rho_min = max(1.4 / max(mat["fy"], 1.0), 0.25 * math.sqrt(max(mat["fc"], 0.0)) / max(mat["fy"], 1.0))
        rho_max = mat["rho_max"]
        fs = float("inf")
        z = 0.0
        a_blk = None
    else:
        a_blk = As * mat["fy"] / (0.85 * mat["fc"] * b)
        if a_blk >= d:
            Mn = 0.0
            phiMn = 0.0
            okM = False
            uM = float("inf")
        else:
            Mn = As * mat["fy"] * (d - 0.5 * a_blk) / 1e6
            phiMn = mat["phi_m"] * Mn
            okM = phiMn >= Mu
            uM = Mu / phiMn if phiMn > 0 else float("inf")

        # ACI-style simplified shear:
        # - With stirrups: Vc ~= 0.17*sqrt(fc)*bw*d (SI)
        # - Without stirrups: use reduced concrete term 0.083*sqrt(fc)*bw*d (SI)
        vc_coef = 0.17 if Asv > 0.0 else 0.083
        Vc = vc_coef * math.sqrt(max(mat["fc"], 0.0)) * b * d / 1000.0
        Vs = Asv * mat["fyv"] * d / max(1.0, cand["s"]) / 1000.0 if Asv > 0.0 else 0.0
        phiVn = mat["phi_v"] * (Vc + Vs)
        okV = phiVn >= Vu
        uV = Vu / phiVn if phiVn > 0 else float("inf")

        rho = As / (b * d)
        rho_min = max(1.4 / max(mat["fy"], 1.0), 0.25 * math.sqrt(max(mat["fc"], 0.0)) / max(mat["fy"], 1.0))
        rho_max = mat["rho_max"]
        okR = rho >= rho_min and rho <= rho_max
        uR = max(rho_min / rho if rho > 0 else float("inf"), rho / rho_max if rho_max > 0 else float("inf"))

        if Asv > 0.0:
            s_lim = min(300.0, 0.75 * d)
            okS = cand["s"] <= s_lim
            uS = cand["s"] / s_lim if s_lim > 0 else float("inf")
        else:
            s_lim = 0.0
            okS = True
            uS = 0.0

        z = 0.9 * d
        fs = Ms * 1e6 / (As * z) if As > 0 and z > 0 else float("inf")
        okFs = fs <= mat["fs_lim"]
        uFs = fs / mat["fs_lim"] if mat["fs_lim"] > 0 else float("inf")

        okD = dmm <= dallow
        uD = dmm / dallow if dallow > 0 else float("inf")

    conc = (b / 1000.0) * (h / 1000.0) * beam["span"]
    form = (2.0 * h / 1000.0 + b / 1000.0) * beam["span"]
    lb = beam["span"] + 2.0 * 40.0 * db / 1000.0
    wt_bot = cand["nb"] * lb * wb
    if cand["nt"] > 0:
        lt = beam["span"] + 2.0 * 30.0 * dt / 1000.0
        wt_top = cand["nt"] * lt * wt
    else:
        wt_top = 0.0
    if Asv > 0.0:
        core_w = max(0.0, b - 2.0 * (mat["cover"] + 0.5 * ds))
        core_h = max(0.0, h - 2.0 * (mat["cover"] + 0.5 * ds))
        l_st = 2.0 * (core_w + core_h) / 1000.0 + 2.0 * 10.0 * ds / 1000.0
        n_st = int(math.floor(beam["span"] * 1000.0 / cand["s"])) + 1
        wt_st = n_st * l_st * ws
    else:
        wt_st = 0.0
    steel = wt_bot + wt_top + wt_st

    cost_jpy = conc * cost["c_conc"] + steel * cost["c_steel"] + form * cost["c_form"]
    co2 = conc * cost["e_conc"] + steel * cost["e_steel"] + form * cost["e_form"]
    mode = setts["mode"]
    obj = cost_jpy if mode == "COST" else (co2 if mode == "CO2" else setts["w_cost"] * cost_jpy + setts["w_co2"] * co2)

    ok = all([okM, okV, okD, okFs, okR, okC, okS])
    ng = []
    if not okM:
        ng.append("flexure")
    if not okV:
        ng.append("shear")
    if not okD:
        ng.append("deflection")
    if not okFs:
        ng.append("steel_stress")
    if not okR:
        ng.append("rho")
    if not okC:
        ng.append("clear_spacing")
    if not okS:
        ng.append("stirrup_spacing")
    if d <= 0:
        ng.insert(0, "d<=0")

    return {
        "beam": beam["id"], "rank": cand["rank"], "sec": cand["sec"], "b": b, "h": h,
        "bot": f"{cand['nb']}-{cand['db']}", "top": (f"{cand['nt']}-{cand['dt']}" if cand["nt"] > 0 else "0"),
        "st": (f"{cand['legs']}L-{cand['ds']}@{cand['s']:g}" if Asv > 0.0 else "NONE"), "Mu": Mu, "phiMn": phiMn, "Vu": Vu, "phiVn": phiVn,
        "dmm": dmm, "dallow": dallow, "fs": fs, "rho": rho, "rho_min": rho_min, "rho_max": rho_max,
        "clear": clear, "clear_req": req_clear, "s": cand["s"], "s_lim": s_lim,
        "conc": conc, "steel": steel, "form": form, "cost": cost_jpy, "co2": co2, "obj": obj,
        "okM": okM, "okV": okV, "okD": okD, "okFs": okFs, "okR": okR, "okC": okC, "okS": okS,
        "ok": ok, "ng": ", ".join(ng), "util": max(uM, uV, uD, uFs, uR, uC, uS),
        "span": beam["span"], "trib": beam["trib"], "qD_area": beam["qD"], "qL_area": beam["qL"],
        "PD": beam["PD"], "PL": beam["PL"], "r": beam["r"], "point_a": a_p,
        "nb": cand["nb"], "db": cand["db"], "nt": cand["nt"], "dt": cand["dt"],
        "legs": cand["legs"], "ds": cand["ds"], "cover": mat["cover"],
        "Ab": Ab, "As": As, "Asv": Asv, "d_eff": d, "a_blk": a_blk, "Mn": Mn,
        "Vc": Vc, "Vs": Vs, "Ec": Ec, "Ig": Ig, "ei": ei, "z": z,
        "sw": sw, "wD": wD, "wL": wL, "wS": wS, "wU": wU, "pS": pS, "pU": pU,
        "R1s": R1s, "R2s": R2s, "R1u": R1u, "R2u": R2u,
        "xMu": xmu, "Ms": Ms, "xMs": xms, "xVu": xvu, "xd": xd,
    }


def optimize(mat, setts, cost, beams, cands, n_override=0):
    checks, best, status = [], {}, {}
    for beam in beams:
        rows = [eval_cand(beam, c, mat, setts, cost, n_override=n_override) for c in cands]
        checks.extend(rows)
        feas = [r for r in rows if r["ok"]]
        if feas:
            sel = min(feas, key=lambda r: (r["obj"], r["rank"], r["util"]))
            status[beam["id"]] = "OK"
        else:
            sel = min(rows, key=lambda r: (r["util"], r["obj"], r["rank"]))
            status[beam["id"]] = "NG"
        best[beam["id"]] = sel
    return checks, best, status


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for c in ws[row]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")


def autosize(ws, maxw=42):
    for col in ws.columns:
        w = 10
        for c in col:
            s = "" if c.value is None else str(c.value)
            w = max(w, min(maxw, len(s) + 2))
        ws.column_dimensions[col[0].column_letter].width = w


def nround(v, digits=3):
    if v is None:
        return None
    try:
        fv = float(v)
    except Exception:
        return v
    if not math.isfinite(fv):
        return None
    return round(fv, digits)


def write_calc_best_beams_sheet(wb, mat, setts, beams, best, status):
    ws = wb.create_sheet("CALC_BEST_BEAMS")
    ws.cell(row=1, column=1, value="CALC_BEST_BEAMS (Selected RC beam calculation sheet)").font = Font(bold=True, size=14)

    selected = [best[b["id"]] for b in beams]
    ok_count = sum(1 for b in beams if status[b["id"]] == "OK")
    ng_count = len(beams) - ok_count
    tconc = sum(r["conc"] for r in selected)
    tsteel = sum(r["steel"] for r in selected)
    tform = sum(r["form"] for r in selected)
    tcost = sum(r["cost"] for r in selected)
    tco2 = sum(r["co2"] for r in selected)
    worst = max(selected, key=lambda r: (float("inf") if not math.isfinite(r["util"]) else r["util"], r["Mu"], r["Vu"], r["dmm"]))

    summary_headers = [
        "BeamCount", "OKCount", "NGCount", "ObjectiveMode", "CostWeight", "CO2Weight",
        "WorstBeamID", "WorstUtil", "TotalCost[JPY]", "TotalCO2[kg]", "Concrete[m3]", "Rebar[kg]", "Formwork[m2]",
    ]
    summary_vals = [
        len(beams), ok_count, ng_count, setts["mode"], setts["w_cost"], setts["w_co2"],
        worst["beam"], nround(worst["util"], 4), nround(tcost, 2), nround(tco2, 3),
        nround(tconc, 4), nround(tsteel, 3), nround(tform, 3),
    ]
    for c, v in enumerate(summary_headers, start=1):
        ws.cell(row=3, column=c, value=v)
    for c, v in enumerate(summary_vals, start=1):
        ws.cell(row=4, column=c, value=v)
    style_header(ws, 3)

    ws.cell(row=6, column=1, value="Strength checks use ULS line loads (wU, pU); deflection uses service loads (wS, pS).").font = Font(italic=True)
    ws.cell(row=7, column=1, value="Array tables below show ULS shear/moment and service deflection on the same 20-division x grid.").font = Font(italic=True)

    ws.cell(row=9, column=1, value="Selected beam summary (one governing candidate per beam)").font = Font(bold=True)
    headers = [
        "No", "BeamID", "Status", "Rank", "Section", "Span[m]", "Trib[m]",
        "qD[kN/m2]", "qL[kN/m2]", "PD[kN]", "PL[kN]", "PointPosRatio", "Point_a[m]",
        "b[mm]", "h[mm]", "Bottom", "Top", "Stirrup",
        "d[mm]", "As[mm2]", "Asv[mm2]", "a_blk[mm]", "Mn[kN*m]", "phiMn[kN*m]",
        "Vc[kN]", "Vs[kN]", "phiVn[kN]", "sw[kN/m]",
        "wD[kN/m]", "wL[kN/m]", "wS[kN/m]", "pS[kN]", "wU[kN/m]", "pU[kN]",
        "R1s[kN]", "R2s[kN]", "R1u[kN]", "R2u[kN]",
        "Mu[kN*m]", "xMu[m]", "Vu[kN]", "xVu[m]",
        "Defl[mm]", "xDefl[m]", "DeflAllow[mm]", "SteelStress[N/mm2]",
        "rho", "Clear[mm]", "Cost[JPY]", "CO2[kg]", "UtilMax", "NG_Reason",
    ]
    for c, v in enumerate(headers, start=1):
        ws.cell(row=10, column=c, value=v)
    style_header(ws, 10)

    row = 11
    for i, beam in enumerate(beams, start=1):
        r = best[beam["id"]]
        ws.append([
            i, beam["id"], status[beam["id"]], r["rank"], r["sec"], nround(r["span"], 3), nround(r["trib"], 3),
            nround(r["qD_area"], 3), nround(r["qL_area"], 3), nround(r["PD"], 3), nround(r["PL"], 3), nround(r["r"], 4), nround(r["point_a"], 3),
            nround(r["b"], 1), nround(r["h"], 1), r["bot"], r["top"], r["st"],
            nround(r["d_eff"], 3), nround(r["As"], 3), nround(r["Asv"], 3), nround(r["a_blk"], 3), nround(r["Mn"], 3), nround(r["phiMn"], 3),
            nround(r["Vc"], 3), nround(r["Vs"], 3), nround(r["phiVn"], 3), nround(r["sw"], 3),
            nround(r["wD"], 3), nround(r["wL"], 3), nround(r["wS"], 3), nround(r["pS"], 3), nround(r["wU"], 3), nround(r["pU"], 3),
            nround(r["R1s"], 3), nround(r["R2s"], 3), nround(r["R1u"], 3), nround(r["R2u"], 3),
            nround(r["Mu"], 3), nround(r["xMu"], 3), nround(r["Vu"], 3), nround(r["xVu"], 3),
            nround(r["dmm"], 3), nround(r["xd"], 3), nround(r["dallow"], 3), nround(r["fs"], 3),
            nround(r["rho"], 6), nround(r["clear"], 3), nround(r["cost"], 2), nround(r["co2"], 3), nround(r["util"], 4), r["ng"],
        ])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=f"Worst selected beam arrays (BeamID={worst['beam']}, Section={worst['sec']}, Util={nround(worst['util'], 4)})").font = Font(bold=True)
    row += 1
    for c, v in enumerate(["x[m]", "V_u[kN]", "M_u[kN*m]", "y_service[mm]"], start=1):
        ws.cell(row=row, column=c, value=v)
    style_header(ws, row)
    row += 1

    xs, vs, ms, ys = response_arrays(worst["span"], worst["wU"], worst["pU"], worst["wS"], worst["pS"], worst["r"], worst["ei"], n=20)
    for i in range(len(xs)):
        ws.append([nround(xs[i], 4), nround(vs[i], 4), nround(ms[i], 4), nround(ys[i], 4)])

    ws.freeze_panes = "A11"


def write_calc_best_beam_arrays_sheet(wb, beams, best, status):
    ws = wb.create_sheet("CALC_BEST_BEAM_ARRAYS")
    ws.cell(row=1, column=1, value="CALC_BEST_BEAM_ARRAYS (selected beam x-V-M-deflection arrays)").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="Each beam is output at fixed 20 divisions. V/M use ULS loads, deflection uses service loads.").font = Font(italic=True)

    row = 4
    for beam in beams:
        r = best[beam["id"]]
        ws.cell(
            row=row,
            column=1,
            value=(
                f"BeamID={beam['id']}, Status={status[beam['id']]}, Span={nround(r['span'], 3)}m, "
                f"Section={r['sec']}(R{r['rank']}), Bottom={r['bot']}, Top={r['top']}, Stirrup={r['st']}"
            ),
        ).font = Font(bold=True)
        row += 1
        for c, v in enumerate(["x[m]", "V_u[kN]", "M_u[kN*m]", "y_service[mm]"], start=1):
            ws.cell(row=row, column=c, value=v)
        style_header(ws, row)
        row += 1

        xs, vs, ms, ys = response_arrays(r["span"], r["wU"], r["pU"], r["wS"], r["pS"], r["r"], r["ei"], n=20)
        for i in range(len(xs)):
            ws.append([nround(xs[i], 4), nround(vs[i], 4), nround(ms[i], 4), nround(ys[i], 4)])
            row += 1
        row += 1


def write_out(out_path, mat, setts, beams, checks, best, status, warnings):
    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("SUMMARY")
    ws2 = wb.create_sheet("CHECKS")

    ws1.append(["BeamID", "Status", "Rank", "Section", "b_mm", "h_mm", "Bottom", "Top", "Stirrup", "Objective", "Cost_JPY", "CO2_kg", "Mu_kNm", "phiMn_kNm", "Vu_kN", "phiVn_kN", "Defl_mm", "DeflAllow_mm", "SteelStress_N/mm2", "UtilMax", "NG_Reason"])
    tconc = tsteel = tform = tcost = tco2 = 0.0
    for b in beams:
        r = best[b["id"]]
        ws1.append([
            b["id"], status[b["id"]], r["rank"], r["sec"], nround(r["b"], 1), nround(r["h"], 1), r["bot"], r["top"], r["st"],
            nround(r["obj"], 2), nround(r["cost"], 2), nround(r["co2"], 3), nround(r["Mu"], 3), nround(r["phiMn"], 3),
            nround(r["Vu"], 3), nround(r["phiVn"], 3), nround(r["dmm"], 3), nround(r["dallow"], 3), nround(r["fs"], 3),
            nround(r["util"], 4), r["ng"],
        ])
        tconc += r["conc"]
        tsteel += r["steel"]
        tform += r["form"]
        tcost += r["cost"]
        tco2 += r["co2"]
    style_header(ws1); ws1.freeze_panes = "A2"

    ws2.append(["BeamID", "Rank", "Section", "b_mm", "h_mm", "Bottom", "Top", "Stirrup", "Mu_kNm", "phiMn_kNm", "Vu_kN", "phiVn_kN", "Defl_mm", "DeflAllow_mm", "SteelStress_N/mm2", "rho", "rho_min", "rho_max", "Clear_mm", "ClearReq_mm", "s_mm", "sLim_mm", "Cost_JPY", "CO2_kg", "Objective", "Flex", "Shear", "Defl", "Stress", "Rho", "Clear", "Stirrup", "Pass", "UtilMax", "NG_Reason"])
    rows = checks
    if not setts["all_checks"]:
        keep = {(v["beam"], v["rank"], v["sec"]) for v in best.values()}
        rows = [r for r in checks if (r["beam"], r["rank"], r["sec"]) in keep]
    if len(rows) > setts["max_rows"]:
        rows = rows[:setts["max_rows"]]
    for r in rows:
        ws2.append([
            r["beam"], r["rank"], r["sec"], nround(r["b"], 1), nround(r["h"], 1), r["bot"], r["top"], r["st"],
            nround(r["Mu"], 3), nround(r["phiMn"], 3), nround(r["Vu"], 3), nround(r["phiVn"], 3),
            nround(r["dmm"], 3), nround(r["dallow"], 3), nround(r["fs"], 3), nround(r["rho"], 6),
            nround(r["rho_min"], 6), nround(r["rho_max"], 6), nround(r["clear"], 3), nround(r["clear_req"], 3),
            nround(r["s"], 3), nround(r["s_lim"], 3), nround(r["cost"], 2), nround(r["co2"], 3), nround(r["obj"], 3),
            "OK" if r["okM"] else "NG", "OK" if r["okV"] else "NG", "OK" if r["okD"] else "NG",
            "OK" if r["okFs"] else "NG", "OK" if r["okR"] else "NG", "OK" if r["okC"] else "NG",
            "OK" if r["okS"] else "NG", "OK" if r["ok"] else "NG", nround(r["util"], 4), r["ng"],
        ])
    style_header(ws2); ws2.freeze_panes = "A2"

    write_calc_best_beams_sheet(wb, mat, setts, beams, best, status)
    write_calc_best_beam_arrays_sheet(wb, beams, best, status)
    ws3 = wb["CALC_BEST_BEAMS"]
    ws4 = wb["CALC_BEST_BEAM_ARRAYS"]

    ws5 = wb.create_sheet("QUANTITY")
    ws5.append(["Item", "Total", "Unit"])
    ws5.append(["Concrete", round(tconc, 4), "m3"])
    ws5.append(["Rebar", round(tsteel, 3), "kg"])
    ws5.append(["Formwork", round(tform, 3), "m2"])
    ws5.append(["Total Cost", round(tcost, 2), "JPY"])
    ws5.append(["Total CO2", round(tco2, 3), "kg-CO2e"])
    ws5.append(["Objective Mode", setts["mode"], "-"])
    ws5.append(["Cost Weight", setts["w_cost"], "-"])
    ws5.append(["CO2 Weight", setts["w_co2"], "-"])
    ws5.append(["fc", mat["fc"], "N/mm2"])
    ws5.append(["fy", mat["fy"], "N/mm2"])
    style_header(ws5)

    ws6 = wb.create_sheet("WARNINGS")
    ws6.append(["Type", "Message"])
    for m in warnings:
        ws6.append(["INPUT_WARNING", m])
    for b in beams:
        if status[b["id"]] != "OK":
            r = best[b["id"]]
            ws6.append(["DESIGN_WARNING", f"{b['id']}: no feasible candidate. closest={r['sec']}(rank={r['rank']}) NG={r['ng']}"])
    style_header(ws6)

    for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
        autosize(ws)
    wb.save(out_path)

def make_template(path):
    wb = Workbook()
    w0 = wb.active
    w0.title = "README"
    w0.append(["RC Beam Optimizer Input Template"])
    w0.append(["For rc_beam_optimizer.py (RC beam version)"])
    w0.append(["Unit system: length=mm,m  force=N,kN  stress=N/mm2  weight=kg"])
    w0.append(["Checks are simplified for screening. Final code check is required."])
    w0.append(["Required sheets: MATERIALS, SETTINGS, COST, BEAMS, CANDIDATES"])
    w0.append(["For no-stirrup case, set StirrupLegs_n=0 and keep StirrupBar blank."])

    w1 = wb.create_sheet("MATERIALS")
    w1.append(["Key", "Value", "Unit/Note"])
    w1.append(["fc_n_mm2", 30, "Concrete compressive strength (N/mm2)"])
    w1.append(["fy_main_n_mm2", 490, "Main bar yield strength (N/mm2)"])
    w1.append(["fy_shear_n_mm2", 295, "Stirrup yield strength (N/mm2)"])
    w1.append(["cover_mm", 40, "Cover"])
    w1.append(["concrete_unit_weight_kn_m3", 24, "Self weight unit"])
    w1.append(["phi_flexure", 0.9, "Phi for flexure"])
    w1.append(["phi_shear", 0.75, "Phi for shear"])
    w1.append(["rho_max", 0.025, "Max rebar ratio"])
    w1.append(["steel_stress_limit_n_mm2", 280, "Service steel stress limit (N/mm2)"])
    w1.append(["deflection_limit_ratio", 250, "L/x"])
    w1.append(["effective_stiffness_factor", 0.35, "EIeff factor"])
    w1.append(["min_clear_spacing_mm", 25, "Min clear spacing"])
    w1.append(["default_n_div", 400, "Numerical division count"])

    w2 = wb.create_sheet("SETTINGS")
    w2.append(["Key", "Value", "Unit/Note"])
    w2.append(["load_factor_dead", 1.2, "ULS dead factor"])
    w2.append(["load_factor_live", 1.6, "ULS live factor"])
    w2.append(["objective_mode", "HYBRID", "COST / CO2 / HYBRID"])
    w2.append(["cost_weight", 1.0, "For HYBRID"])
    w2.append(["co2_weight", 0.002, "For HYBRID"])
    w2.append(["max_check_rows", 20000, "CHECKS cap"])
    w2.append(["output_all_checks", True, "TRUE/FALSE"])

    w3 = wb.create_sheet("COST")
    w3.append(["Key", "Value", "Unit/Note"])
    w3.append(["concrete_jpy_m3", 18500, "JPY/m3"])
    w3.append(["rebar_jpy_kg", 165, "JPY/kg"])
    w3.append(["formwork_jpy_m2", 4200, "JPY/m2"])
    w3.append(["concrete_co2_kg_m3", 320, "kg-CO2e/m3"])
    w3.append(["rebar_co2_kg_kg", 1.4, "kg-CO2e/kg"])
    w3.append(["formwork_co2_kg_m2", 8.0, "kg-CO2e/m2"])

    w4 = wb.create_sheet("BEAMS")
    w4.append(["Use", "BeamID", "Span_m", "TributaryWidth_m", "DeadLoad_kN_m2", "LiveLoad_kN_m2", "PointDead_kN", "PointLive_kN", "PointPosRatio", "Notes"])
    w4.append([True, "B1", 6.0, 3.0, 4.0, 3.0, 0.0, 0.0, 0.50, "Typical floor beam"])
    w4.append([True, "B2", 7.2, 2.8, 4.5, 3.5, 90.0, 30.0, 0.50, "Machine load at midspan"])
    w4.append([True, "B3", 5.4, 2.5, 5.0, 2.0, 0.0, 0.0, 0.50, "Short span"])
    w4.append([True, "B4", 8.0, 3.2, 4.0, 3.0, 60.0, 20.0, 0.40, "Off-center point load"])

    w5 = wb.create_sheet("CANDIDATES")
    w5.append(["Use", "Rank", "SectionID", "Width_mm", "Depth_mm", "BottomBars_n", "BottomBar", "TopBars_n", "TopBar", "StirrupLegs_n", "StirrupBar", "StirrupSpacing_mm"])
    w5.append([True, 1, "C300x500_A", 300, 500, 3, "D22", 2, "D13", 2, "D10", 150])
    w5.append([True, 2, "C300x550_A", 300, 550, 3, "D22", 2, "D13", 2, "D10", 150])
    w5.append([True, 3, "C300x550_B", 300, 550, 4, "D22", 2, "D13", 2, "D10", 150])
    w5.append([True, 4, "C350x550_A", 350, 550, 4, "D22", 2, "D16", 2, "D10", 175])
    w5.append([True, 5, "C350x600_A", 350, 600, 4, "D22", 2, "D16", 2, "D13", 175])
    w5.append([True, 6, "C350x600_B", 350, 600, 5, "D22", 2, "D16", 2, "D13", 150])
    w5.append([True, 7, "C400x600_A", 400, 600, 4, "D25", 2, "D16", 2, "D13", 175])
    w5.append([True, 8, "C400x650_A", 400, 650, 4, "D25", 2, "D16", 2, "D13", 150])
    w5.append([True, 9, "C400x650_B", 400, 650, 5, "D25", 2, "D16", 2, "D13", 150])
    w5.append([True, 10, "C450x700_A", 450, 700, 5, "D25", 2, "D19", 2, "D13", 150])
    w5.append([True, 11, "C450x700_B", 450, 700, 6, "D25", 2, "D19", 2, "D13", 150])

    for ws in wb.worksheets:
        style_header(ws)
        autosize(ws)
    wb.save(path)


def main(argv=None):
    p = argparse.ArgumentParser(description="RC beam optimizer")
    p.add_argument("input", nargs="?", default="input_rc_beam.xlsx")
    p.add_argument("output", nargs="?", default="output_rc_beam.xlsx")
    p.add_argument("--make-template", dest="template", default="")
    p.add_argument("--n-div", dest="n_div", type=int, default=0)
    a = p.parse_args(sys.argv[1:] if argv is None else argv)

    if a.template:
        make_template(a.template)
        print(f"Template created: {a.template}")
        return 0

    if not os.path.exists(a.input):
        print(f"Input file not found: {a.input}", file=sys.stderr)
        print("Tip: use --make-template", file=sys.stderr)
        return 2

    try:
        mat, setts, cost, beams, cands, warnings = read_input(a.input)
        checks, best, status = optimize(mat, setts, cost, beams, cands, n_override=(a.n_div if a.n_div > 0 else 0))
        write_out(a.output, mat, setts, beams, checks, best, status, warnings)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    ok = sum(1 for b in beams if status[b["id"]] == "OK")
    print(f"Done. Output: {a.output}")
    print(f"Beams: {len(beams)}  OK: {ok}  NG: {len(beams)-ok}")
    if warnings:
        print(f"Input warnings: {len(warnings)} (see WARNINGS sheet)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
